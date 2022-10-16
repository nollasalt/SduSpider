import requests
import re
import openpyxl
import datetime
import os

url='https://www.bkjx.sdu.edu.cn/sanji_list.jsp?totalpage=9&PAGENUM=1&urltype=tree.TreeTempUrl&wbtreeid=1013'
headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'}
page_data=requests.get(url=url,headers=headers).text
#获取第一页的title与link
links='div style="float:left"><a href="(.*?)"'
titles='target=_blank title="(.*?)" style=""'
times='<div style="float:right;">(.*?)</div>'
links_data=re.findall(links,page_data,re.S)
titles_data=re.findall(titles,page_data,re.S)
times_data=re.findall(times,page_data,re.S)

#获取总页数    
pagemax=int(re.findall('TD nowrap align=left width=1% id=fanye128813>共130条&nbsp;&nbsp;1/(.*?)&nbsp',page_data,re.S)[0])
#获取所有title与link
for i in range(2,pagemax+1):
    url='https://www.bkjx.sdu.edu.cn/sanji_list.jsp?totalpage=9&PAGENUM='+str(i)+'&urltype=tree.TreeTempUrl&wbtreeid=1013'
    page_data=requests.get(url=url,headers=headers).text
    links_data.extend(re.findall(links,page_data,re.S))
    titles_data.extend(re.findall(titles,page_data,re.S))
    times_data.extend(re.findall(times,page_data,re.S))
#对不完整的链接进行修正
for i in range(len(links_data)):
    if re.findall('(.*?)content',links_data[i],re.S)==['']:
        links_data[i]='https://www.bkjx.sdu.edu.cn/'+links_data[i]

print(len(times_data))
print(len(titles_data))
print(len(links_data))

ti = datetime.datetime.now()
time_hot = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时' + str(ti.minute) + '分'
time_file = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时'
# 文件保存位置
file_path = './山大新闻' + time_file + '.xlsx'     # 以小时为单位生成excel表
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()        # 创建excel表格
    wb.save(file_path)
wb = openpyxl.load_workbook(file_path)
ws = wb.create_sheet(title=time_hot)    # 创建数据表，以时间命名
ws.cell(row=1, column=1, value='序号')    # 表头
ws.cell(row=1, column=2, value='链接')
ws.cell(row=1, column=3, value='标题')
ws.cell(row=1, column=4, value='时间')
for i in range(len(links_data)):
    ws.cell(row=i+2, column=1, value=i+1)   # 热度排名
    ws.cell(row=i+2, column=3, value=titles_data[i])      # 问题内容
    ws.cell(row=i+2, column=2, value=links_data[i])
    ws.cell(row=i+2, column=4, value=times_data[i])
    # print('第', i+1, '个问题保存成功')
if 'Sheet' in wb.sheetnames:        # 删除自动生成的’Sheet‘表格
    del wb['Sheet']
wb.save(file_path)      # 保存并关闭文件
wb.close()



